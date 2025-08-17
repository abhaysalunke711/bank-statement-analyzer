"""
Excel viewer utility for displaying Excel content in web browser.
Converts Excel sheets to HTML tables with interactive features.
"""

import os
import logging
from typing import Dict, List, Any
import openpyxl
from openpyxl import load_workbook
import json

class ExcelViewer:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def read_excel_file(self, excel_path: str) -> Dict[str, Any]:
        """
        Read Excel file and convert to web-displayable format.
        
        Args:
            excel_path: Path to Excel file
            
        Returns:
            Dictionary with sheet data and metadata
        """
        if not os.path.exists(excel_path):
            self.logger.error(f"Excel file not found: {excel_path}")
            return {}
        
        try:
            workbook = load_workbook(excel_path)
            sheets_data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = self._convert_sheet_to_data(sheet, sheet_name)
                sheets_data[sheet_name] = sheet_data
            
            workbook.close()
            
            return {
                'filename': os.path.basename(excel_path),
                'sheets': sheets_data,
                'sheet_names': list(workbook.sheetnames),
                'total_sheets': len(workbook.sheetnames)
            }
            
        except Exception as e:
            self.logger.error(f"Error reading Excel file {excel_path}: {e}")
            return {}
    
    def _convert_sheet_to_data(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Convert Excel sheet to structured data."""
        try:
            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            if max_row == 1 and max_col == 1:
                return {
                    'name': sheet_name,
                    'headers': [],
                    'rows': [],
                    'summary': {'total_rows': 0, 'total_cols': 0},
                    'is_empty': True
                }
            
            # Extract data
            data = []
            for row in sheet.iter_rows(values_only=True):
                # Convert None values to empty strings and handle different data types
                cleaned_row = []
                for cell in row:
                    if cell is None:
                        cleaned_row.append('')
                    elif isinstance(cell, (int, float)):
                        cleaned_row.append(str(cell))
                    else:
                        cleaned_row.append(str(cell))
                data.append(cleaned_row)
            
            # Remove empty rows from the end
            while data and all(cell == '' or cell is None for cell in data[-1]):
                data.pop()
            
            if not data:
                return {
                    'name': sheet_name,
                    'headers': [],
                    'rows': [],
                    'summary': {'total_rows': 0, 'total_cols': 0},
                    'is_empty': True
                }
            
            # Determine headers and data rows
            headers = data[0] if data else []
            rows = data[1:] if len(data) > 1 else []
            
            # Clean up headers
            headers = [str(h) if h else f'Column {i+1}' for i, h in enumerate(headers)]
            
            # Detect sheet type and add metadata
            sheet_type = self._detect_sheet_type(sheet_name, headers, rows)
            
            return {
                'name': sheet_name,
                'type': sheet_type,
                'headers': headers,
                'rows': rows,
                'summary': {
                    'total_rows': len(rows),
                    'total_cols': len(headers),
                    'has_data': len(rows) > 0
                },
                'is_empty': False
            }
            
        except Exception as e:
            self.logger.error(f"Error processing sheet {sheet_name}: {e}")
            return {
                'name': sheet_name,
                'headers': [],
                'rows': [],
                'summary': {'total_rows': 0, 'total_cols': 0, 'error': str(e)},
                'is_empty': True
            }
    
    def _detect_sheet_type(self, sheet_name: str, headers: List[str], rows: List[List[str]]) -> str:
        """Detect the type of sheet based on name and content."""
        sheet_name_lower = sheet_name.lower()
        headers_lower = [h.lower() for h in headers]
        
        # Check for pivot-style layout
        if ('financial summary' in sheet_name_lower or 
            (len(headers) > 3 and any(month in ' '.join(headers_lower) for month in 
             ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']) and
             'total' in ' '.join(headers_lower))):
            return 'pivot'
        elif 'summary' in sheet_name_lower:
            return 'summary'
        elif any(month in sheet_name_lower for month in ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 
                                                        'jul', 'aug', 'sep', 'oct', 'nov', 'dec']):
            return 'monthly'
        elif 'transaction' in ' '.join(headers_lower):
            return 'transactions'
        elif 'category' in ' '.join(headers_lower):
            return 'categories'
        else:
            return 'data'
    
    def generate_html_tables(self, excel_data: Dict[str, Any]) -> str:
        """
        Generate HTML tables from Excel data.
        
        Args:
            excel_data: Excel data from read_excel_file
            
        Returns:
            HTML string with interactive tables
        """
        if not excel_data or not excel_data.get('sheets'):
            return "<p class='text-muted'>No data to display</p>"
        
        html_parts = []
        
        # Add tabs navigation
        html_parts.append(self._generate_tabs_nav(excel_data))
        
        # Add tab content
        html_parts.append('<div class="tab-content" id="excelTabContent">')
        
        for i, (sheet_name, sheet_data) in enumerate(excel_data['sheets'].items()):
            active_class = 'show active' if i == 0 else ''
            html_parts.append(f'''
                <div class="tab-pane fade {active_class}" id="sheet-{i}" role="tabpanel">
                    <div class="card mt-3">
                        <div class="card-header">
                            <h5 class="mb-0">
                                <i class="fas fa-table me-2"></i>
                                {sheet_name}
                                <span class="badge bg-info ms-2">{sheet_data['summary']['total_rows']} rows</span>
                            </h5>
                        </div>
                        <div class="card-body">
                            {self._generate_sheet_table(sheet_data)}
                        </div>
                    </div>
                </div>
            ''')
        
        html_parts.append('</div>')
        
        return '\n'.join(html_parts)
    
    def _generate_tabs_nav(self, excel_data: Dict[str, Any]) -> str:
        """Generate tabs navigation HTML."""
        nav_parts = ['<ul class="nav nav-tabs" id="excelTabs" role="tablist">']
        
        for i, sheet_name in enumerate(excel_data['sheet_names']):
            active_class = 'active' if i == 0 else ''
            sheet_data = excel_data['sheets'][sheet_name]
            sheet_type = sheet_data.get('type', 'data')
            
            # Choose icon based on sheet type
            icon = {
                'pivot': 'fas fa-table',
                'summary': 'fas fa-chart-pie',
                'monthly': 'fas fa-calendar-alt',
                'transactions': 'fas fa-list',
                'categories': 'fas fa-tags',
                'data': 'fas fa-table'
            }.get(sheet_type, 'fas fa-table')
            
            nav_parts.append(f'''
                <li class="nav-item" role="presentation">
                    <button class="nav-link {active_class}" id="tab-{i}" data-bs-toggle="tab" 
                            data-bs-target="#sheet-{i}" type="button" role="tab">
                        <i class="{icon} me-1"></i>
                        {sheet_name}
                        <span class="badge bg-secondary ms-1">{sheet_data['summary']['total_rows']}</span>
                    </button>
                </li>
            ''')
        
        nav_parts.append('</ul>')
        return '\n'.join(nav_parts)
    
    def _generate_sheet_table(self, sheet_data: Dict[str, Any]) -> str:
        """Generate HTML table for a single sheet."""
        if sheet_data['is_empty'] or not sheet_data['rows']:
            return '<p class="text-muted">No data in this sheet</p>'
        
        # Add summary info
        summary_html = f'''
            <div class="row mb-3">
                <div class="col-md-6">
                    <small class="text-muted">
                        <i class="fas fa-info-circle me-1"></i>
                        {sheet_data['summary']['total_rows']} rows × {sheet_data['summary']['total_cols']} columns
                    </small>
                </div>
                <div class="col-md-6 text-end">
                    <small class="text-muted">
                        <i class="fas fa-layer-group me-1"></i>
                        Sheet type: {sheet_data.get('type', 'data').title()}
                    </small>
                </div>
            </div>
        '''
        
        # Generate table
        table_html = ['<div class="table-responsive">']
        table_html.append('<table class="table table-striped table-hover">')
        
        # Headers
        table_html.append('<thead class="table-dark">')
        table_html.append('<tr>')
        for header in sheet_data['headers']:
            table_html.append(f'<th scope="col">{header}</th>')
        table_html.append('</tr>')
        table_html.append('</thead>')
        
        # Body
        table_html.append('<tbody>')
        for row_idx, row in enumerate(sheet_data['rows']):
            # Add zebra striping and hover effects
            row_class = 'table-light' if row_idx % 2 == 0 else ''
            table_html.append(f'<tr class="{row_class}">')
            
            for col_idx, cell in enumerate(row):
                # Format cells based on content with row context for better color coding
                header = sheet_data['headers'][col_idx] if col_idx < len(sheet_data['headers']) else ''
                formatted_cell = self._format_cell(cell, header, row)
                table_html.append(f'<td>{formatted_cell}</td>')
            
            table_html.append('</tr>')
        table_html.append('</tbody>')
        table_html.append('</table>')
        table_html.append('</div>')
        
        return summary_html + '\n'.join(table_html)
    
    def _format_cell(self, cell: str, header: str, row_data: List[str] = None) -> str:
        """Format cell content based on type and header with income/expense color coding."""
        if not cell or cell == '':
            return '<span class="text-muted">—</span>'
        
        header_lower = header.lower()
        
        # Detect transaction type from row data for color coding
        transaction_type = None
        if row_data:
            # Look for 'Type' column in the row
            try:
                type_index = next(i for i, h in enumerate(['Type', 'Transaction Type', 'Income/Expense']) 
                                if any(keyword in h for keyword in ['type', 'income', 'expense']))
                if type_index < len(row_data):
                    transaction_type = row_data[type_index].lower()
            except (StopIteration, IndexError):
                # Try to infer from amount if available
                if 'amount' in header_lower:
                    try:
                        amount = float(str(cell).replace('$', '').replace(',', '').replace('(', '-').replace(')', ''))
                        transaction_type = 'income' if amount > 0 else 'expense'
                    except (ValueError, TypeError):
                        pass
        
        # Format currency amounts with income/expense colors
        if 'amount' in header_lower or 'total' in header_lower or '$' in str(cell):
            try:
                # Try to parse as currency
                if '$' in str(cell) or ',' in str(cell):
                    amount_str = str(cell)
                    # Apply color based on transaction type or amount sign
                    if transaction_type == 'income' or (not transaction_type and not '-' in amount_str and not '(' in amount_str):
                        return f'<span class="text-end fw-bold text-success bg-light-success px-2 py-1 rounded">{cell}</span>'
                    else:
                        return f'<span class="text-end fw-bold text-danger bg-light-danger px-2 py-1 rounded">{cell}</span>'
                else:
                    # Try to parse as number
                    num = float(cell)
                    if num < 0:
                        return f'<span class="text-end fw-bold text-danger bg-light-danger px-2 py-1 rounded">${abs(num):,.2f}</span>'
                    else:
                        return f'<span class="text-end fw-bold text-success bg-light-success px-2 py-1 rounded">${num:,.2f}</span>'
            except (ValueError, TypeError):
                pass
        
        # Format transaction type with color coding
        if 'type' in header_lower and cell.lower() in ['income', 'expense']:
            if cell.lower() == 'income':
                return f'<span class="badge bg-success"><i class="fas fa-arrow-up me-1"></i>{cell}</span>'
            else:
                return f'<span class="badge bg-danger"><i class="fas fa-arrow-down me-1"></i>{cell}</span>'
        
        # Format confidence percentages
        if 'confidence' in header_lower and '%' in str(cell):
            try:
                conf_val = float(str(cell).replace('%', ''))
                if conf_val >= 80:
                    return f'<span class="badge bg-success">{cell}</span>'
                elif conf_val >= 60:
                    return f'<span class="badge bg-warning">{cell}</span>'
                else:
                    return f'<span class="badge bg-secondary">{cell}</span>'
            except (ValueError, TypeError):
                pass
        
        # Format dates
        if 'date' in header_lower and '/' in str(cell):
            return f'<span class="text-nowrap">{cell}</span>'
        
        # Format categories with badges and special handling for pivot layout
        if 'category' in header_lower and cell != 'Uncategorized':
            # Special formatting for pivot section headers
            if cell in ['💰 INCOME', '💸 EXPENSES']:
                if '💰' in cell:
                    return f'<span class="badge bg-success fs-6 py-2 px-3"><i class="fas fa-arrow-up me-2"></i>{cell}</span>'
                else:
                    return f'<span class="badge bg-danger fs-6 py-2 px-3"><i class="fas fa-arrow-down me-2"></i>{cell}</span>'
            
            # Special formatting for total rows
            if 'TOTAL' in cell.upper():
                if 'INCOME' in cell.upper():
                    return f'<span class="badge bg-success fw-bold py-2 px-3">{cell}</span>'
                elif 'EXPENSE' in cell.upper():
                    return f'<span class="badge bg-danger fw-bold py-2 px-3">{cell}</span>'
                elif 'NET' in cell.upper():
                    return f'<span class="badge bg-primary fw-bold py-2 px-3">{cell}</span>'
            
            badge_colors = {
                'Food & Dining': 'bg-warning',
                'Shopping': 'bg-info',
                'Transportation': 'bg-primary',
                'Entertainment': 'bg-success',
                'Utilities': 'bg-secondary',
                'Healthcare': 'bg-danger',
                'Banking & Finance': 'bg-dark',
                'Salary': 'bg-success',
                'Investment': 'bg-info',
                'Government': 'bg-primary',
                'Business Income': 'bg-success'
            }
            color = badge_colors.get(cell, 'bg-light text-dark')
            return f'<span class="badge {color}">{cell}</span>'
        
        # Format descriptions with transaction type coloring
        if 'description' in header_lower and transaction_type:
            if transaction_type == 'income':
                return f'<span class="text-success-emphasis">{cell}</span>'
            elif transaction_type == 'expense':
                return f'<span class="text-danger-emphasis">{cell}</span>'
        
        # Long descriptions - truncate with tooltip
        if len(str(cell)) > 50:
            truncated = str(cell)[:50] + '...'
            return f'<span title="{cell}" data-bs-toggle="tooltip">{truncated}</span>'
        
        return str(cell)
    
    def get_summary_stats(self, excel_data: Dict[str, Any]) -> Dict[str, Any]:
        """Extract summary statistics from Excel data."""
        if not excel_data or not excel_data.get('sheets'):
            return {}
        
        stats = {
            'total_sheets': excel_data['total_sheets'],
            'sheets_info': []
        }
        
        for sheet_name, sheet_data in excel_data['sheets'].items():
            sheet_info = {
                'name': sheet_name,
                'type': sheet_data.get('type', 'data'),
                'rows': sheet_data['summary']['total_rows'],
                'cols': sheet_data['summary']['total_cols'],
                'has_data': sheet_data['summary']['has_data']
            }
            
            # Extract financial data if available
            if sheet_data.get('type') == 'summary':
                sheet_info['is_summary'] = True
                # Try to extract key financial metrics
                for row in sheet_data['rows']:
                    if len(row) >= 2:
                        key = str(row[0]).lower()
                        value = row[1]
                        if 'total income' in key:
                            sheet_info['total_income'] = value
                        elif 'total expenses' in key:
                            sheet_info['total_expenses'] = value
                        elif 'net amount' in key:
                            sheet_info['net_amount'] = value
            
            stats['sheets_info'].append(sheet_info)
        
        return stats
